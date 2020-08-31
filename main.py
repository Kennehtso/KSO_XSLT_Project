# -*- coding: utf-8 -*-
"""
Created on Sun Aug 30 16:11:35 2020
Provide a pratice to merge data from different excel table 
@author: kenne
"""
 
import openpyxl

class Subjects(object):
    def __init__(self, name, interval):
        self.name = name
        self.interval = interval
        
class Carers(object):
    def __init__(self, name, timeStamp):
        self.name = name
        self.timeStamp = timeStamp

def getAvailableCarers(time, carers):
    resultList = []
    for i in range(0, len(carers)):
        name = carers[i].name
        if time in carers[i].timeStamp:
            resultList.append(carers[i].name)
    if len(resultList) == 0:
        resultList.append('--找不到關懷員匹配--')
    return resultList

def frange(x, y, jump):  
  while x < y:  
    yield x  
    x += jump 
    
def getSubjectsData(sheet, timePeriod, carers, wb):
    resultList = []
    # generate new result sheet
    wb.create_sheet('Merge_Result',len(wb.worksheets))
    sheet_Merge_Result = wb.worksheets[len(wb.worksheets)-1]
    
    for i in range(1, sheet_Subjects.max_row+1):
        currentSubjectnName = sheet_Subjects.cell(row=i, column=1).value
        #UPDATE RESULT SHEET - ROW TITLE
        sheet_Merge_Result.cell(row=i, column=1).value = currentSubjectnName
        
        interval = {}
        timePeriodIndex = 0
        for j in range (2, sheet_Subjects.max_column + 1):
        #for j in range(2, len(timePeriod)) :
            timeStamp = sheet_Subjects.cell(row=i, column=j).value
            if timeStamp != None:
                interval[timeStamp] = getAvailableCarers(timeStamp, carers)
                sheet_Merge_Result.cell(row=i, column=j).value = str(interval[timeStamp])
                #subject = Subjects(currentSubjectnName,timeStamp, getAvailableCarers(timeStamp, carers))
            else:
                interval[timePeriod[timePeriodIndex]] = ['--此時段不適用--']
                sheet_Merge_Result.cell(row=i, column=j).value = str(interval[timePeriod[timePeriodIndex]])
               #subject = Subjects(currentSubjectnName,timeStamp, ['--此時段不適用--']) 
            timePeriodIndex+=1
        subject = Subjects(currentSubjectnName, interval) 
        resultList.append(subject)
    
    
    sheet_Merge_Result.insert_rows(1)
    index = 0
    for i in range(2, sheet_Merge_Result.max_column + 1):
         sheet_Merge_Result.cell(row=1, column=i).value = timePeriod[index]
         index+=1
    # Log
    for i in range(0, len(resultList)):
        print(resultList[i].name,' : ')
        for item in resultList[i].interval.items():
            print(item)
            print("----------------------------")
        print("-------------------------------------------------------------")
    
    
    wb.save("Carer Schedule.xlsx")
    #return resultList


def getCarersData(sheet):
    resultList = []
    for i in range(1, sheet.max_row+1):
        currenCarerName = sheet.cell(row=i, column=1).value
        carers =  Carers(currenCarerName, [])
        timestamps = []
        for j in range (2, sheet.max_column+1):
            val = sheet.cell(row=i, column=j).value
            timestamps.append(val)
        carers.timeStamp = timestamps
        # Append to whole Subjects List
        resultList.append(carers)
        
    # Log
    for i in range(0, len(resultList)):
        print(resultList[i].name,';',resultList[i].timeStamp)
        
    return resultList

def getMaxValue(sheet):
    for i in range (1, sheet.max_row+1):
        val = sheet.cell(row=i, column= sheet.max_column).value
        if val != None:
            print("max value is : ", val)
            return val
        
def getTimePeriod(maxValue):
    resultList = []
    v = 1.1
    ref = 1
    while v < maxValue:
        if ref % 5!=0:
            resultList.append(round(v,1))
        v += 0.1
        ref += 1
          
    # Log
    print('Full List ',resultList)
    return resultList


""" Create xlsx Sample 
wb = openpyxl.Workbook()
sheet = wb.worksheets[0]
sheet['A1'] = 'Hello'
listTitle = ['Name','Phone']
sheet.append(listTitle)
wb.save('test.xlsx')
"""

""" Read & Merge data"""
wb = openpyxl.load_workbook("Carer Schedule.xlsx")
sheet_Carers = wb.worksheets[1]
sheet_Subjects = wb.worksheets[0]

maxValue = getMaxValue(sheet_Subjects)
timePeriod = getTimePeriod(maxValue) 

carers = getCarersData(sheet_Carers)

#Update Result By subjects
subjects = getSubjectsData(sheet_Subjects, timePeriod, carers ,wb)



""" End fo R&M """